from openpyxl import load_workbook
import json

if __name__ == '__main__':
    #Save ground truth code to translation_results.xlsx
    #50 datapoints each time because target languish is different
    with open('/test_functions.json') as infile:
        funcs = json.load(infile)
    lang = "Java"
    wb = load_workbook('/translation_results.xlsx')
    sh = wb["Sheet1"]
    for i in range(1, sh.max_row + 1):
        key = str(sh.cell(i, 1).value)
        # print(key)
        # print(funcs[lang][key])
        sh.cell(i, 3).value = funcs[lang][key]
    wb.save('/translation_results.xlsx')


    #Save all prediction and ground truth datas into 2 txt files to get pre file and ref file
    p = []
    q = []
    wb = load_workbook('./results.xlsx')
    sh = wb["Sheet1"]
    for i in range(1, sh.max_row + 1):
        p.append(sh.cell(i, 1).value.strip().replace('\n', ''))
        # q.append(sh.cell(i,2).value.strip().replace('\n', ''))
    print(p)
    file = open("/dev_output.txt", 'w')
    for i in range(0, len(p)):
        file.write(p[i] + '\n')
    file.close()
    #pre and ref files are save in this project

    #evaluate
    #!python drive/MyDrive/Code-Code/code-to-code-trans/evaluator/CodeBLEU/calc_code_bleu.py --refs /content/drive/MyDrive/text_ref.txt --hyp /content/drive/MyDrive/text_output.txt --lang java --params 0.25,0.25,0.25,0.25